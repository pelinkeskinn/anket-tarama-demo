from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.db import Base, get_session
from app.main import app
from app.models import AnswerResult, StoredFormCreate
from app.database import create_form
from app.scoring import SCORE_MAP


def _override_db(tmp_path):  # type: ignore[no-untyped-def]
    engine = create_engine(f"sqlite:///{(tmp_path / 'export.db').as_posix()}")
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    def override() -> Session:  # type: ignore[misc]
        with TestingSession() as session:
            yield session

    app.dependency_overrides[get_session] = override
    return TestingSession


def _answers() -> list[AnswerResult]:
    values = ["NEVER", "SOMETIMES", "ALWAYS", "BLANK"]
    statuses = ["MARKED", "MARKED", "MARKED", "BLANK"]
    answers = []
    for question_no in range(1, 16):
        if question_no == 6:
            answers.append(
                AnswerResult(questionNo=6, value=None, confidence=0.4, source="UNRESOLVED", status="UNCERTAIN")
            )
            continue
        index = (question_no - 1) % 4
        answers.append(
            AnswerResult(
                questionNo=question_no,
                value=values[index],  # type: ignore[arg-type]
                confidence=0.9,
                source="AUTO",
                status=statuses[index],  # type: ignore[arg-type]
            )
        )
    return answers


def test_numeric_export_scores_and_legend(tmp_path) -> None:  # type: ignore[no-untyped-def]
    TestingSession = _override_db(tmp_path)
    with TestingSession() as session:
        create_form(
            StoredFormCreate(
                analysisId="export-1",
                templateCode="HEALTHY_NUTRITION_V3",
                formConfidence=0.9,
                answers=_answers(),
            ),
            session,
        )

    client = TestClient(app)
    try:
        response = client.get("/api/forms/export.xlsx?format=numeric")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["Anket Kayıtları"]
        assert "Özet" not in workbook.sheetnames
        key_sheet = workbook["Puan Anahtarı"]
        assert sheet["F1"].value == "Soru 1"
        assert sheet["Q1"].value == "Soru 12"
        assert sheet["F2"].value == SCORE_MAP["NEVER"]
        assert sheet["G2"].value == SCORE_MAP["SOMETIMES"]
        assert sheet["H2"].value == SCORE_MAP["ALWAYS"]
        assert sheet["I2"].value is None
        assert sheet["K2"].value is None
        fill = str(sheet["K2"].fill.fgColor.rgb or sheet["K2"].fill.fgColor.theme)
        assert "F4A261" in fill.upper() or sheet["K2"].fill.patternType == "solid"
        header_values = [cell.value for cell in sheet[1]]
        assert "Güven (%)" not in header_values
        assert "Toplam Puan" not in header_values
        assert sheet.cell(1, 21).value == "Yanıtlanan Soru Sayısı"
        assert isinstance(sheet.cell(2, 21).value, (int, float))
        assert key_sheet["A1"].value == "Puan Anahtarı"
        assert key_sheet["A4"].value == "Soru 1-15"
        assert key_sheet["A6"].value == "Hiçbir zaman"
        assert key_sheet["B6"].value == int(SCORE_MAP["NEVER"])
        assert key_sheet["A7"].value == "Bazen"
        assert key_sheet["B7"].value == int(SCORE_MAP["SOMETIMES"])
        assert key_sheet["A8"].value == "Her zaman"
        legend_text = " ".join(str(cell.value) for row in key_sheet.iter_rows() for cell in row if cell.value)
        assert "NEVER" not in legend_text
        assert "ALWAYS" not in legend_text
        assert "SOMETIMES" not in legend_text
        assert "OFTEN" not in legend_text
        text_response = client.get("/api/forms/export.xlsx?format=text")
        text_book = load_workbook(BytesIO(text_response.content))
        assert text_book.active["F2"].value == "Hiçbir zaman"
        assert "Puan Anahtarı" in text_book.sheetnames
        questions_sheet = text_book["Anket Soruları"]
        assert questions_sheet["A16"].value == 15
        assert "Tuzlu gıdaları" in questions_sheet["B2"].value
        assert "Özet" not in text_book.sheetnames
    finally:
        app.dependency_overrides.clear()


def test_forms_list_is_paginated(tmp_path) -> None:  # type: ignore[no-untyped-def]
    TestingSession = _override_db(tmp_path)
    with TestingSession() as session:
        for index in range(3):
            create_form(
                StoredFormCreate(
                    analysisId=f"page-{index}",
                        templateCode="HEALTHY_NUTRITION_V3",
                    formConfidence=0.9,
                    answers=_answers(),
                ),
                session,
            )
    client = TestClient(app)
    try:
        response = client.get("/api/forms?limit=2&offset=0")
        payload = response.json()
        assert response.status_code == 200
        assert payload["total"] == 3
        assert len(payload["items"]) == 2
        assert "possibleDuplicate" in payload["items"][0]
    finally:
        app.dependency_overrides.clear()


def test_admin_token_protects_export_and_delete(tmp_path, monkeypatch) -> None:  # type: ignore[no-untyped-def]
    monkeypatch.setenv("ADMIN_TOKEN", "test-admin")
    _override_db(tmp_path)
    client = TestClient(app)
    try:
        listed = client.get("/api/forms")
        assert listed.status_code == 200
        export_denied = client.get("/api/forms/export.xlsx")
        assert export_denied.status_code == 401
        delete_denied = client.delete("/api/forms/1")
        assert delete_denied.status_code == 401
    finally:
        app.dependency_overrides.clear()
