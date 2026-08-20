from __future__ import annotations

from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.auth import require_admin
from app.database import create_form, delete_form, get_form, iter_form_details, list_forms
from app.db import get_session
from app.models import StoredFormCreate, StoredFormDetail, StoredFormPage
from app.scoring import FREQUENCY_QUESTION_START, SCORE_MAP, answer_score, is_review_cell


router = APIRouter(prefix="/api/forms", tags=["forms"])

REVIEW_FILL = PatternFill("solid", fgColor="F4A261")
MAX_EXPORT_QUESTIONS = 26
LEGEND_TEMPLATE = "HEALTHY_NUTRITION_V2"
EMPTY_SCORE_NOTE = "Boş hücre — sayıma dahil edilmez"


def _not_found() -> HTTPException:
    return HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Kayıt bulunamadı."}})


@router.post("", response_model=StoredFormDetail, status_code=201)
def save_form(payload: StoredFormCreate, session: Session = Depends(get_session)) -> StoredFormDetail:
    return create_form(payload, session)


@router.get("", response_model=StoredFormPage)
def forms(
    session: Session = Depends(get_session),
    _: None = Depends(require_admin),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
) -> StoredFormPage:
    items, total = list_forms(session, limit=limit, offset=offset)
    return StoredFormPage(items=items, total=total)


@router.get("/export.xlsx")
def export_forms(
    session: Session = Depends(get_session),
    _: None = Depends(require_admin),
    format: Literal["numeric", "text"] = Query(default="numeric"),
) -> StreamingResponse:
    forms = list(iter_form_details(session))
    question_count = max(
        [MAX_EXPORT_QUESTIONS, *[max((answer.questionNo for answer in form.answers), default=0) for form in forms]]
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anket Kayıtları"
    headers = [
        "Kayıt No",
        "Tarih",
        "Şablon",
        "Güven (%)",
        "Boş",
        "Manuel",
        *[
            f"Soru {number} ({'G' if number < FREQUENCY_QUESTION_START else 'S'})"
            for number in range(1, question_count + 1)
        ],
        "Toplam Puan",
        "Yanıtlanan Soru Sayısı",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="C1121F")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for form in forms:
        answers_by_no = {answer.questionNo: answer for answer in form.answers}
        scored_values: list[float] = []
        answered = 0
        row_values: list[object] = [
            form.id,
            form.createdAt.replace(tzinfo=None),
            form.templateCode,
            round(form.formConfidence * 100, 1),
            form.blankCount,
            form.manualCount,
        ]
        review_columns: list[int] = []
        for number in range(1, question_count + 1):
            answer = answers_by_no.get(number)
            value = answer.value if answer else None
            status = answer.status if answer else None
            if format == "text":
                row_values.append(_answer_label_text(form.templateCode, number, value))
                score = answer_score(value, status)
            else:
                score = answer_score(value, status)
                row_values.append(score)
            if score is not None:
                scored_values.append(score)
                answered += 1
            elif is_review_cell(value, status):
                review_columns.append(6 + number)
        total_score = sum(scored_values)
        row_values.extend([total_score, answered])
        sheet.append(row_values)
        if format == "numeric":
            for column in review_columns:
                sheet.cell(sheet.max_row, column).fill = REVIEW_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 21
    sheet.column_dimensions["C"].width = 25
    last_column = 8 + question_count
    for column in range(4, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 17

    _write_score_key_sheet(workbook, header_fill, question_count)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "anket-kayitlari.xlsx" if format == "numeric" else "anket-kayitlari-metin.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _write_score_key_sheet(workbook: Workbook, header_fill: PatternFill, question_count: int) -> None:
    sheet = workbook.create_sheet("Puan Anahtarı")
    sheet["A1"] = "Puan Anahtarı"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A1:B1")
    sheet["A2"] = "G = genel sorular, S = sıklık soruları. Belirsiz hücreler 1. sayfada turuncu işaretlenir."
    sheet.merge_cells("A2:B2")

    general_end = min(FREQUENCY_QUESTION_START - 1, question_count)
    frequency_end = max(question_count, FREQUENCY_QUESTION_START)
    _append_legend_table(
        sheet,
        start_row=4,
        title=f"Genel sorular (Soru 1-{general_end})",
        sample_question=1,
        header_fill=header_fill,
    )
    _append_legend_table(
        sheet,
        start_row=13,
        title=f"Sıklık soruları (Soru {FREQUENCY_QUESTION_START}-{frequency_end})",
        sample_question=FREQUENCY_QUESTION_START,
        header_fill=header_fill,
    )
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 42


def _append_legend_table(
    sheet,
    *,
    start_row: int,
    title: str,
    sample_question: int,
    header_fill: PatternFill,
) -> None:
    sheet.cell(start_row, 1, title).font = Font(bold=True)
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    header_row = start_row + 1
    sheet.cell(header_row, 1, "Görünen cevap")
    sheet.cell(header_row, 2, "Puan")
    for column in (1, 2):
        cell = sheet.cell(header_row, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    row = header_row + 1
    for value in SCORE_MAP:
        sheet.cell(row, 1, _answer_label_text(LEGEND_TEMPLATE, sample_question, value))
        score = SCORE_MAP[value]
        sheet.cell(row, 2, int(score) if score == int(score) else score)
        row += 1
    sheet.cell(row, 1, "Boş")
    sheet.cell(row, 2, EMPTY_SCORE_NOTE)
    row += 1
    sheet.cell(row, 1, "Belirsiz")
    sheet.cell(row, 2, f"{EMPTY_SCORE_NOTE} (turuncu)")


def _answer_label_text(template_code: str, question_no: int, value: str | None) -> str:
    if value is None:
        return "Belirsiz"
    if value == "BLANK":
        return "Boş"
    if template_code.startswith("HEALTHY_NUTRITION_V"):
        if question_no >= 12:
            return {"NEVER": "Hiçbir zaman", "SOMETIMES": "1-2 kez/hafta", "OFTEN": "3-4 kez/hafta", "ALWAYS": "5+ kez/hafta"}.get(value, value)
        return {"NEVER": "Hiçbir zaman", "SOMETIMES": "Ara sıra", "OFTEN": "Sık sık", "ALWAYS": "Her zaman"}.get(value, value)
    return {"NEVER": "Hiçbir zaman", "SOMETIMES": "Bazen", "OFTEN": "Sık sık", "ALWAYS": "Her zaman"}.get(value, value)


_answer_label = _answer_label_text


@router.get("/{form_id}", response_model=StoredFormDetail)
def form_detail(form_id: int, session: Session = Depends(get_session), _: None = Depends(require_admin)) -> StoredFormDetail:
    form = get_form(form_id, session)
    if form is None:
        raise _not_found()
    return form


@router.delete("/{form_id}")
def remove_form(form_id: int, session: Session = Depends(get_session), _: None = Depends(require_admin)) -> dict[str, bool]:
    if not delete_form(form_id, session):
        raise _not_found()
    return {"deleted": True}


# Keep SCORE_MAP import visible for tests that patch or inspect this module.
__all__ = ["SCORE_MAP", "export_forms", "_answer_label", "_answer_label_text"]
